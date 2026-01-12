"""
Exporter 서비스
매핑된 데이터를 Excel/CSV로 내보내기
"""
import os
import io
import uuid
from datetime import datetime
from typing import Optional
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models.schemas import FieldMapping, ExportFormat, ObjectType
from app.models.salesmap import get_object_name, FIELD_TYPE_NORMALIZERS


EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')


@dataclass
class ExportResult:
    """내보내기 결과"""
    success: bool
    filename: str
    file_path: str
    stats: dict
    errors: list[str]


class DataExporter:
    """데이터 내보내기 서비스"""

    def __init__(self, exports_dir: str = EXPORTS_DIR):
        self.exports_dir = exports_dir
        os.makedirs(exports_dir, exist_ok=True)

    def export(
        self,
        data: list[dict],
        mappings: list[FieldMapping],
        object_types: list[str],
        format: ExportFormat = ExportFormat.EXCEL,
        include_summary: bool = True,
    ) -> ExportResult:
        """
        데이터 내보내기

        Args:
            data: 원본 데이터
            mappings: 필드 매핑
            object_types: 오브젝트 타입
            format: 내보내기 형식
            include_summary: 요약 시트 포함 여부

        Returns:
            ExportResult
        """
        errors = []
        stats = {
            "total_rows": len(data),
            "object_counts": {},
        }

        try:
            # 오브젝트별로 데이터 분리
            object_data = self._split_by_object(data, mappings, object_types)

            for obj_type, obj_rows in object_data.items():
                stats["object_counts"][obj_type] = len(obj_rows)

            # 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            unique_id = str(uuid.uuid4())[:8]

            if format == ExportFormat.EXCEL:
                filename = f"salesmap_import_{timestamp}_{unique_id}.xlsx"
                file_path = os.path.join(self.exports_dir, filename)
                self._export_excel(object_data, mappings, file_path, include_summary)
            else:
                # CSV는 첫 번째 오브젝트만 내보내기 (멀티시트 불가)
                first_obj = list(object_data.keys())[0] if object_data else "data"
                filename = f"salesmap_{first_obj}_{timestamp}_{unique_id}.csv"
                file_path = os.path.join(self.exports_dir, filename)
                self._export_csv(object_data.get(first_obj, []), file_path)

            return ExportResult(
                success=True,
                filename=filename,
                file_path=file_path,
                stats=stats,
                errors=errors,
            )

        except Exception as e:
            return ExportResult(
                success=False,
                filename="",
                file_path="",
                stats=stats,
                errors=[str(e)],
            )

    def _split_by_object(
        self,
        data: list[dict],
        mappings: list[FieldMapping],
        object_types: list[str],
    ) -> dict[str, list[dict]]:
        """오브젝트별로 데이터 분리"""
        result = {obj: [] for obj in object_types}

        # 오브젝트별 매핑 그룹화
        obj_mappings = {}
        for m in mappings:
            obj = m.target_object
            if obj not in obj_mappings:
                obj_mappings[obj] = []
            obj_mappings[obj].append(m)

        for row in data:
            for obj_type in object_types:
                obj_row = {}
                obj_maps = obj_mappings.get(obj_type, [])

                for m in obj_maps:
                    source_value = row.get(m.source_column)
                    # 값 정규화
                    normalized = self._normalize_value(source_value, m.field_type)
                    # 필드 라벨을 컬럼명으로 사용
                    obj_row[m.target_field_label] = normalized

                if obj_row:  # 빈 행 제외
                    result[obj_type].append(obj_row)

        return result

    def _normalize_value(self, value, field_type: str):
        """필드 타입에 맞게 값 정규화"""
        if value is None or str(value).strip() == '':
            return ''

        normalizer = FIELD_TYPE_NORMALIZERS.get(field_type)
        if normalizer:
            try:
                return normalizer(value)
            except (ValueError, TypeError):
                return str(value)

        return str(value)

    def _export_excel(
        self,
        object_data: dict[str, list[dict]],
        mappings: list[FieldMapping],
        file_path: str,
        include_summary: bool,
    ):
        """Excel 파일로 내보내기"""
        wb = Workbook()

        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 요약 시트 추가
        if include_summary:
            self._add_summary_sheet(wb, object_data, mappings)

        # 오브젝트별 시트 추가
        for obj_type, rows in object_data.items():
            if not rows:
                continue

            obj_name = get_object_name(obj_type)
            sheet = wb.create_sheet(title=obj_name)

            # DataFrame으로 변환
            df = pd.DataFrame(rows)

            # 헤더 스타일
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 데이터 작성
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

                    if r_idx == 1:  # 헤더
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')

            # 컬럼 너비 자동 조정
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def _add_summary_sheet(
        self,
        wb: Workbook,
        object_data: dict[str, list[dict]],
        mappings: list[FieldMapping],
    ):
        """요약 시트 추가"""
        sheet = wb.create_sheet(title="요약", index=0)

        # 스타일 정의
        title_font = Font(size=14, bold=True)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # 제목
        sheet['A1'] = "세일즈맵 데이터 이관 요약"
        sheet['A1'].font = title_font
        sheet.merge_cells('A1:C1')

        # 생성 정보
        sheet['A3'] = "생성일시"
        sheet['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 오브젝트별 행 수
        sheet['A5'] = "오브젝트별 데이터"
        sheet['A5'].font = header_font

        row = 6
        for obj_type, rows in object_data.items():
            obj_name = get_object_name(obj_type)
            sheet[f'A{row}'] = obj_name
            sheet[f'B{row}'] = f"{len(rows)}행"
            row += 1

        # 필드 매핑 정보
        row += 1
        sheet[f'A{row}'] = "필드 매핑"
        sheet[f'A{row}'].font = header_font

        row += 1
        sheet[f'A{row}'] = "원본 컬럼"
        sheet[f'B{row}'] = "대상 필드"
        sheet[f'C{row}'] = "타입"
        sheet[f'A{row}'].fill = header_fill
        sheet[f'B{row}'].fill = header_fill
        sheet[f'C{row}'].fill = header_fill

        for m in mappings:
            row += 1
            sheet[f'A{row}'] = m.source_column
            sheet[f'B{row}'] = m.target_field_label
            # Enum인 경우 .value 사용, 아니면 그대로
            field_type_str = m.field_type.value if hasattr(m.field_type, 'value') else str(m.field_type)
            sheet[f'C{row}'] = field_type_str

        # 컬럼 너비
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 15

    def _export_csv(self, rows: list[dict], file_path: str):
        """CSV 파일로 내보내기"""
        if not rows:
            with open(file_path, 'w', encoding='utf-8-sig') as f:
                f.write('')
            return

        df = pd.DataFrame(rows)
        df.to_csv(file_path, index=False, encoding='utf-8-sig')


# 싱글톤 인스턴스
exporter = DataExporter()
